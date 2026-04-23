#!/usr/bin/env python3
"""Convert Navicat MySQL backup files into struct and data SQL files."""

from __future__ import annotations

import argparse
import glob
import re
import shlex
import shutil
import socket
import subprocess
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font

PROGRAM_NAME = "convertnavbak"
PROGRAM_FILE = "convertnavbak_v1.13.py"
DEFAULT_CONFIG_FILENAME = "convertnavbak.conf"
DEFAULT_PROC = 0
DEFAULT_TOT_ROWS = 500
DEFAULT_INSERT_PREFIX = "ins_"
DEFAULT_DB_PORT = 3306
CONFIG_KEYS = [
    "input",
    "proc",
    "one_struct",
    "tot_rows",
    "insert_prefix",
    "db_host",
    "db_user",
    "db_pw",
    "db_port",
    "db_name",
]

ONE_STRUCT_ON = {"y", "1", "true", "on", "yes"}
ONE_STRUCT_OFF = {"", "false", "off", "0", "no", "n"}

NAVICAT_HEADER_RE = re.compile(
    r"Navicat(?:\s+Premium)?\s+(?:Dump\s+SQL|Data\s+Transfer)",
    flags=re.IGNORECASE,
)
MYSQL_SOURCE_RE = re.compile(
    r"Source\s+Server\s+Type\s*:\s*MySQL",
    flags=re.IGNORECASE,
)
STRUCT_RE = re.compile(r"(?is)^\s*(DROP\s+TABLE(?:\s+IF\s+EXISTS)?|CREATE\s+TABLE)\b")
INSERT_RE = re.compile(r"(?is)^\s*INSERT\s+INTO\b")
ANY_INSERT_RE = re.compile(r"\bINSERT\s+INTO\b", flags=re.IGNORECASE)

_INTEGER_AND_DATE_TYPE_RE = re.compile(
    r"(?i)\b(?:tinyint|smallint|mediumint|int|integer|bigint|date|datetime|timestamp|time|year)\s*\(\s*\d+\s*\)"
)
_CURRENT_TIMESTAMP_PRECISION_RE = re.compile(
    r"(?i)\b(?:CURRENT_TIMESTAMP|NOW|LOCALTIME|LOCALTIMESTAMP)\s*\(\s*\d+\s*\)"
)

_LAST_STATUS_WIDTH = 0


@dataclass
class FileResult:
    path: Path
    ti: str = ""
    original_insert_count: int = 0
    new_insert_count: int = 0
    duration_seconds: float = 0.0
    start_dt: Optional[datetime] = None
    end_dt: Optional[datetime] = None
    status: str = "S"
    error: str = ""
    struct_outputs: List[Path] = field(default_factory=list)
    data_output: Optional[Path] = None


class SpecError(Exception):
    """Raised for specification-related validation or parse errors."""


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog=PROGRAM_NAME,
        description="Process Navicat MySQL backup files and split them into struct and data SQL outputs.",
        add_help=False,
        allow_abbrev=False,
    )
    parser.add_argument("input", nargs="?", help="Input file, folder, or wildcard pattern.")
    parser.add_argument("-help", "-h", action="store_true", dest="show_help", help="Show this help message and exit.")
    parser.add_argument(
        "-proc",
        dest="proc",
        nargs="?",
        const=str(DEFAULT_PROC),
        default=None,
        help=(
            "Processing mode. 0 = convert SQL files then optionally generate SQL Batch Upload Script File; "
            "1 = convert SQL files only; 2 = generate SQL Batch Upload Script File only."
        ),
    )
    parser.add_argument(
        "-one-struct",
        "-ost",
        dest="one_struct",
        nargs="?",
        const="",
        default=None,
        help=(
            "When set to on/true/1/yes/y, write all struct statements into one combined file. "
            "When omitted, empty, false, 0, off, or no, write one struct file per input file."
        ),
    )
    parser.add_argument(
        "-tot_rows",
        "-tr",
        dest="tot_rows",
        nargs="?",
        const=str(DEFAULT_TOT_ROWS),
        default=None,
        help=f"Maximum rows per generated INSERT statement. Default: {DEFAULT_TOT_ROWS}.",
    )
    parser.add_argument(
        "-insert_prefix",
        "-inspref",
        dest="insert_prefix",
        nargs="?",
        const=DEFAULT_INSERT_PREFIX,
        default=None,
        help=(
            "Prefix to prepend to the generated INSERT SQL filename. "
            f"Default: {DEFAULT_INSERT_PREFIX}."
        ),
    )
    parser.add_argument("-db_host", dest="db_host", default=None, help="MySQL db host name, url, or ip address.")
    parser.add_argument("-db_user", dest="db_user", default=None, help="MySQL db username.")
    parser.add_argument("-db_pw", dest="db_pw", default=None, help="MySQL db user password.")
    parser.add_argument("-db_port", dest="db_port", default=None, help=f"MySQL port number, defaults to {DEFAULT_DB_PORT}.")
    parser.add_argument("-db_name", dest="db_name", default=None, help="MySQL target database name where to upload the data.")
    parser.add_argument("-config", dest="config", default=None, help=f"Config filename. Default: {DEFAULT_CONFIG_FILENAME}.")
    return parser

def print_examples(parser: argparse.ArgumentParser) -> None:
    parser.print_help()
    print()
    print("Examples:")
    print(f'  python {PROGRAM_FILE} "/path/to/file.sql"')
    print(f'  python {PROGRAM_FILE} "/path/to/folder" -proc=1 -tr=500 -inspref=ins_')
    print(f'  python {PROGRAM_FILE} "/path/to/*.sql" -ost=y -tr=1000')
    print(f'  python {PROGRAM_FILE} -proc=2 -config=myjob.conf')

def default_config_path() -> Path:
    return Path.cwd() / DEFAULT_CONFIG_FILENAME


def user_provided_cli_parameters(args: argparse.Namespace) -> bool:
    values = [
        args.input,
        args.proc,
        args.one_struct,
        args.tot_rows,
        args.insert_prefix,
        args.db_host,
        args.db_user,
        args.db_pw,
        args.db_port,
        args.db_name,
        args.config,
    ]
    return any(value is not None for value in values)

def read_config_file(path: Path) -> Dict[str, str]:
    config: Dict[str, str] = {}
    if not path.exists():
        return config

    for raw_line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or line.startswith(";"):
            continue
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if key in CONFIG_KEYS:
            config[key] = value.strip()
    return config


def normalize_config_values(config: Dict[str, str]) -> Dict[str, str]:
    normalized = {key: config.get(key, "") for key in CONFIG_KEYS}
    if not normalized["proc"]:
        normalized["proc"] = str(DEFAULT_PROC)
    if not normalized["tot_rows"]:
        normalized["tot_rows"] = str(DEFAULT_TOT_ROWS)
    if not normalized["insert_prefix"]:
        normalized["insert_prefix"] = DEFAULT_INSERT_PREFIX
    if not normalized["db_port"]:
        normalized["db_port"] = str(DEFAULT_DB_PORT)
    return normalized

def write_config_file(path: Path, config: Dict[str, str]) -> None:
    normalized = normalize_config_values(config)
    lines = [f"{key}={normalized.get(key, '')}" for key in CONFIG_KEYS]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")


def apply_cli_overrides(config: Dict[str, str], args: argparse.Namespace) -> Dict[str, str]:
    updated = dict(config)
    mapping = {
        "input": args.input,
        "proc": args.proc,
        "one_struct": args.one_struct,
        "tot_rows": args.tot_rows,
        "insert_prefix": args.insert_prefix,
        "db_host": args.db_host,
        "db_user": args.db_user,
        "db_pw": args.db_pw,
        "db_port": args.db_port,
        "db_name": args.db_name,
    }
    for key, value in mapping.items():
        if value is not None:
            updated[key] = str(value)
    return updated

def display_config_file(path: Path, config: Dict[str, str]) -> None:
    normalized = normalize_config_values(config)
    write_config_file(path, normalized)
    print()
    print(f"Config file to use: {path}")
    print("-" * 60)
    print(path.read_text(encoding="utf-8", errors="replace"), end="")
    print("-" * 60)


def prompt_text(message: str) -> str:
    try:
        return input(message)
    except EOFError as exc:
        raise SpecError("Interactive input ended before confirmation was completed.") from exc


def prompt_yes_no(message: str) -> bool:
    while True:
        answer = prompt_text(message).strip().lower()
        if answer in {"y", "yes"}:
            return True
        if answer in {"n", "no"}:
            return False
        print("Please answer Y/Yes or N/No.")


PROMPT_LABELS = {
    "input": "Input file, folder, or wildcard pattern",
    "proc": "-proc value",
    "one_struct": "-one-struct / -ost value",
    "tot_rows": "-tot_rows / -tr value",
    "insert_prefix": "-insert_prefix / -inspref value",
    "db_host": "-db_host value",
    "db_user": "-db_user value",
    "db_pw": "-db_pw value",
    "db_port": "-db_port value",
    "db_name": "-db_name value",
}


def prompt_for_config_values(config: Dict[str, str]) -> Dict[str, str]:
    updated = dict(normalize_config_values(config))
    for key in CONFIG_KEYS:
        current_value = updated.get(key, "")
        entered_value = prompt_text(f"{PROMPT_LABELS[key]} [{current_value}]: ").strip()
        if entered_value:
            updated[key] = entered_value
    return updated


def resolve_active_config(args: argparse.Namespace) -> Tuple[Path, Dict[str, str]]:
    default_path = default_config_path()
    cli_has_parameters = user_provided_cli_parameters(args)
    user_defined_config_provided = args.config is not None

    if user_defined_config_provided:
        config_path = Path(args.config)
        config_values = read_config_file(config_path)
        config_values = apply_cli_overrides(config_values, args)
        write_config_file(config_path, config_values)
        return config_path, normalize_config_values(config_values)

    if cli_has_parameters:
        config_path = default_path
        config_values = read_config_file(config_path)
        config_values = apply_cli_overrides(config_values, args)
        write_config_file(config_path, config_values)
        return config_path, normalize_config_values(config_values)

    if default_path.exists():
        return default_path, normalize_config_values(read_config_file(default_path))

    raise SpecError("No input parameters were provided and the default config file does not exist.")


def review_and_confirm_config(path: Path, config: Dict[str, str]) -> Dict[str, str]:
    current = normalize_config_values(config)
    while True:
        display_config_file(path, current)
        if prompt_yes_no("Review config and proceed? [Y/N]: "):
            write_config_file(path, current)
            return current
        current = prompt_for_config_values(current)


def parse_one_struct(value: Optional[str]) -> bool:
    if value is None:
        return False
    normalized = value.strip().lower()
    if normalized in ONE_STRUCT_OFF:
        return False
    if normalized in ONE_STRUCT_ON:
        return True
    raise SpecError(
        "Invalid value for -one-struct / -ost. Use one of: "
        "Y, y, 1, true, ON, on, YES, TRUE, false, OFF, off, 0, No, N, or omit the value."
    )


def parse_proc(value: Optional[str]) -> int:
    raw_value = str(DEFAULT_PROC) if value in {None, ""} else str(value).strip()
    try:
        parsed = int(raw_value)
    except ValueError as exc:
        raise SpecError("-proc must be 0, 1, or 2.") from exc
    if parsed not in {0, 1, 2}:
        raise SpecError("-proc must be 0, 1, or 2.")
    return parsed

def parse_tot_rows(value: Optional[str]) -> int:
    raw_value = str(DEFAULT_TOT_ROWS) if value in {None, ""} else str(value).strip()
    try:
        parsed = int(raw_value)
    except ValueError as exc:
        raise SpecError("-tot_rows / -tr must be a whole number.") from exc
    if parsed <= 0:
        raise SpecError("-tot_rows / -tr must be greater than zero.")
    return parsed


def parse_insert_prefix(value: Optional[str]) -> str:
    prefix = DEFAULT_INSERT_PREFIX if value in {None, ""} else str(value).strip()
    if not prefix:
        raise SpecError("-insert_prefix / -inspref must not be empty.")
    if "/" in prefix or "\\" in prefix:
        raise SpecError("-insert_prefix / -inspref must not contain path separators.")
    return prefix


def parse_db_port(value: Optional[str]) -> int:
    raw_value = str(DEFAULT_DB_PORT) if value in {None, ""} else str(value).strip()
    try:
        parsed = int(raw_value)
    except ValueError as exc:
        raise SpecError("-db_port must be a whole number.") from exc
    if parsed <= 0:
        raise SpecError("-db_port must be greater than zero.")
    return parsed


def resolve_input_paths(input_expr: str) -> List[Path]:
    input_path = Path(input_expr)

    if input_path.exists():
        if input_path.is_file():
            return [input_path]
        if input_path.is_dir():
            return sorted(
                [p for p in input_path.iterdir() if p.is_file() and p.suffix.lower() == ".sql"],
                key=lambda p: (p.name.lower(), str(p).lower()),
            )
        return []

    matches = [Path(p) for p in glob.glob(input_expr)]
    files = [p for p in matches if p.is_file() and p.suffix.lower() == ".sql"]
    return sorted(files, key=lambda p: (p.name.lower(), str(p).lower()))


def determine_shared_output_dir(files: Sequence[Path]) -> Path:
    if not files:
        return Path.cwd()
    parents = {p.resolve().parent for p in files}
    if len(parents) == 1:
        return next(iter(parents))
    return Path.cwd()


def timestamp_for_filename(dt: Optional[datetime] = None) -> str:
    dt = dt or datetime.now()
    return dt.strftime("%y%m%d_%I%M%S%p").lower()


def timestamp_for_log_rotation(dt: Optional[datetime] = None) -> str:
    dt = dt or datetime.now()
    return dt.strftime("%Y%m%d_%I%M%S%p").lower()


def rotate_existing_log(log_path: Path, when: Optional[datetime] = None) -> Optional[Path]:
    if not log_path.exists():
        return None
    stem = log_path.stem
    suffix = log_path.suffix
    rotated = log_path.with_name(f"{stem}_{timestamp_for_log_rotation(when)}{suffix}")
    counter = 1
    while rotated.exists():
        rotated = log_path.with_name(f"{stem}_{timestamp_for_log_rotation(when)}_{counter}{suffix}")
        counter += 1
    log_path.rename(rotated)
    return rotated


def time_for_log(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    return dt.strftime("%I:%M:%S%p").lower()


def duration_for_log(seconds: float) -> str:
    total_centiseconds = max(0, int(round(seconds * 100)))
    hours = total_centiseconds // 360000
    minutes = (total_centiseconds % 360000) // 6000
    centiseconds = total_centiseconds % 6000
    return f"{hours:02d}.{minutes:02d}.{centiseconds:04d}"


def number_for_log(value: int) -> str:
    return f"{value:,}"


def compute_progress_percentage(
    total_existing_insert_statements: int,
    processed_existing_insert_statements: int,
) -> float:
    if total_existing_insert_statements <= 0:
        return 0.0
    percentage = (processed_existing_insert_statements / total_existing_insert_statements) * 100.0
    return min(100.0, percentage)


def format_status_line(
    file_name: str,
    total_existing_insert_statements: int,
    running_total_new_insert_statements: int,
    processed_existing_insert_statements: int,
    error: str = "",
) -> str:
    percentage = compute_progress_percentage(
        total_existing_insert_statements,
        processed_existing_insert_statements,
    )
    line = (
        f"{file_name} | existing insert: {total_existing_insert_statements} | "
        f"new insert: {running_total_new_insert_statements} | "
        f"percent: {percentage:.2f}%"
    )
    if error:
        line += f" | error: {error}"
    return line


def display_status_line(message: str, done: bool = False) -> None:
    global _LAST_STATUS_WIDTH
    padded = message
    if len(message) < _LAST_STATUS_WIDTH:
        padded = message + (" " * (_LAST_STATUS_WIDTH - len(message)))
    _LAST_STATUS_WIDTH = max(_LAST_STATUS_WIDTH, len(message))
    sys.stdout.write("\r" + padded)
    if done:
        sys.stdout.write("\n")
        _LAST_STATUS_WIDTH = 0
    sys.stdout.flush()


def validate_sql_file_header(text: str, path: Path) -> List[str]:
    errors: List[str] = []
    if path.suffix.lower() != ".sql":
        errors.append("file extension must be .sql")

    header_present = NAVICAT_HEADER_RE.search(text) is not None
    if header_present and MYSQL_SOURCE_RE.search(text) is None:
        errors.append("header indicates Navicat export but is missing 'Source Server Type : MySQL'")

    return errors


def split_sql_statements(sql_text: str) -> List[str]:
    """Split SQL text into statements, ignoring comments and preserving quoted text."""
    statements: List[str] = []
    buf: List[str] = []
    i = 0
    n = len(sql_text)
    in_single = False
    in_double = False
    in_backtick = False
    in_line_comment = False
    in_block_comment = False

    while i < n:
        ch = sql_text[i]
        nxt = sql_text[i + 1] if i + 1 < n else ""
        nxt2 = sql_text[i + 2] if i + 2 < n else ""

        if in_line_comment:
            if ch == "\n":
                in_line_comment = False
            i += 1
            continue

        if in_block_comment:
            if ch == "*" and nxt == "/":
                in_block_comment = False
                i += 2
            else:
                i += 1
            continue

        if in_single:
            buf.append(ch)
            if ch == "\\" and i + 1 < n:
                buf.append(sql_text[i + 1])
                i += 2
                continue
            if ch == "'":
                if nxt == "'":
                    buf.append(nxt)
                    i += 2
                    continue
                in_single = False
            i += 1
            continue

        if in_double:
            buf.append(ch)
            if ch == "\\" and i + 1 < n:
                buf.append(sql_text[i + 1])
                i += 2
                continue
            if ch == '"':
                if nxt == '"':
                    buf.append(nxt)
                    i += 2
                    continue
                in_double = False
            i += 1
            continue

        if in_backtick:
            buf.append(ch)
            if ch == "`":
                in_backtick = False
            i += 1
            continue

        if ch == "-" and nxt == "-" and (not nxt2 or nxt2.isspace()):
            in_line_comment = True
            i += 2
            continue
        if ch == "#":
            in_line_comment = True
            i += 1
            continue
        if ch == "/" and nxt == "*":
            in_block_comment = True
            i += 2
            continue

        if ch == "'":
            in_single = True
            buf.append(ch)
            i += 1
            continue
        if ch == '"':
            in_double = True
            buf.append(ch)
            i += 1
            continue
        if ch == "`":
            in_backtick = True
            buf.append(ch)
            i += 1
            continue

        if ch == ";":
            statement = "".join(buf).strip()
            if statement:
                statements.append(statement)
            buf = []
            i += 1
            continue

        buf.append(ch)
        i += 1

    tail = "".join(buf).strip()
    if tail:
        statements.append(tail)

    return statements


def find_keyword_outside_quotes(text: str, keyword: str) -> int:
    keyword_lower = keyword.lower()
    n = len(text)
    k_len = len(keyword)
    i = 0
    in_single = False
    in_double = False
    in_backtick = False

    while i < n:
        ch = text[i]
        nxt = text[i + 1] if i + 1 < n else ""

        if in_single:
            if ch == "\\" and i + 1 < n:
                i += 2
                continue
            if ch == "'":
                if nxt == "'":
                    i += 2
                    continue
                in_single = False
            i += 1
            continue

        if in_double:
            if ch == "\\" and i + 1 < n:
                i += 2
                continue
            if ch == '"':
                if nxt == '"':
                    i += 2
                    continue
                in_double = False
            i += 1
            continue

        if in_backtick:
            if ch == "`":
                in_backtick = False
            i += 1
            continue

        if ch == "'":
            in_single = True
            i += 1
            continue
        if ch == '"':
            in_double = True
            i += 1
            continue
        if ch == "`":
            in_backtick = True
            i += 1
            continue

        if text[i : i + k_len].lower() == keyword_lower:
            prev_ok = i == 0 or not (text[i - 1].isalnum() or text[i - 1] == "_")
            next_index = i + k_len
            next_ok = next_index >= n or not (text[next_index].isalnum() or text[next_index] == "_")
            if prev_ok and next_ok:
                return i

        i += 1

    return -1


def split_insert_value_groups(values_part: str) -> List[str]:
    groups: List[str] = []
    buf: List[str] = []
    depth = 0
    i = 0
    n = len(values_part)
    in_single = False
    in_double = False
    in_backtick = False

    while i < n:
        ch = values_part[i]
        nxt = values_part[i + 1] if i + 1 < n else ""

        if in_single:
            buf.append(ch)
            if ch == "\\" and i + 1 < n:
                buf.append(values_part[i + 1])
                i += 2
                continue
            if ch == "'":
                if nxt == "'":
                    buf.append(nxt)
                    i += 2
                    continue
                in_single = False
            i += 1
            continue

        if in_double:
            buf.append(ch)
            if ch == "\\" and i + 1 < n:
                buf.append(values_part[i + 1])
                i += 2
                continue
            if ch == '"':
                if nxt == '"':
                    buf.append(nxt)
                    i += 2
                    continue
                in_double = False
            i += 1
            continue

        if in_backtick:
            buf.append(ch)
            if ch == "`":
                in_backtick = False
            i += 1
            continue

        if ch == "'":
            in_single = True
            buf.append(ch)
            i += 1
            continue
        if ch == '"':
            in_double = True
            buf.append(ch)
            i += 1
            continue
        if ch == "`":
            in_backtick = True
            buf.append(ch)
            i += 1
            continue

        if ch == "(":
            depth += 1
            buf.append(ch)
            i += 1
            continue
        if ch == ")":
            if depth == 0:
                raise SpecError("unbalanced parentheses in INSERT values")
            depth -= 1
            buf.append(ch)
            i += 1
            continue

        if ch == "," and depth == 0:
            group = "".join(buf).strip()
            if group:
                groups.append(group)
            buf = []
            i += 1
            continue

        buf.append(ch)
        i += 1

    if depth != 0 or in_single or in_double or in_backtick:
        raise SpecError("unterminated INSERT values content")

    tail = "".join(buf).strip()
    if tail:
        groups.append(tail)

    normalized_groups = [group for group in groups if group]
    for group in normalized_groups:
        if not group.startswith("(") or not group.endswith(")"):
            raise SpecError("malformed INSERT value group")

    return normalized_groups


def parse_insert_statement(statement: str) -> Tuple[str, List[str]]:
    stripped = statement.strip()
    if not INSERT_RE.match(stripped):
        raise SpecError("statement is not an INSERT INTO statement")

    values_index = find_keyword_outside_quotes(stripped, "VALUES")
    if values_index < 0:
        raise SpecError("INSERT statement does not contain VALUES")

    prefix = stripped[: values_index + len("VALUES")].strip()
    values_part = stripped[values_index + len("VALUES") :].strip()
    groups = split_insert_value_groups(values_part)
    if not groups:
        raise SpecError("INSERT statement has no value groups")
    return prefix, groups


def extract_insert_target_name(statement: str) -> str:
    stripped = statement.strip()
    match = re.match(
        r"(?is)^\s*INSERT\s+INTO\s+((?:`[^`]+`|[A-Za-z0-9_]+)(?:\s*\.\s*(?:`[^`]+`|[A-Za-z0-9_]+))?)",
        stripped,
    )
    if not match:
        raise SpecError("unable to determine INSERT target table")
    return re.sub(r"\s+", "", match.group(1))


def detect_single_insert_target(statements: Sequence[str]) -> str:
    targets = {extract_insert_target_name(statement) for statement in statements if INSERT_RE.match(statement)}
    if not targets:
        raise SpecError("unable to determine INSERT target table")
    if len(targets) != 1:
        raise SpecError("each generated insert file must contain only one table")
    return next(iter(targets))

def regroup_insert_statements(
    statements: Sequence[str],
    max_rows: int,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> List[str]:
    if max_rows <= 0:
        raise SpecError("-tot_rows / -tr must be greater than zero")

    emitted: List[str] = []
    current_prefix: Optional[str] = None
    current_groups: List[str] = []
    processed_existing_insert_statements = 0

    def flush() -> None:
        nonlocal current_prefix, current_groups
        if current_prefix is None:
            current_groups = []
            return
        for start in range(0, len(current_groups), max_rows):
            chunk = current_groups[start : start + max_rows]
            emitted.append(f"{current_prefix} {', '.join(chunk)};")
        current_prefix = None
        current_groups = []

    for statement in statements:
        prefix, groups = parse_insert_statement(statement)
        processed_existing_insert_statements += 1
        if current_prefix is None:
            current_prefix = prefix
            current_groups = list(groups)
        elif prefix != current_prefix:
            flush()
            current_prefix = prefix
            current_groups = list(groups)
        else:
            current_groups.extend(groups)

        if progress_callback is not None:
            progress_callback(len(emitted), processed_existing_insert_statements)

    flush()
    if progress_callback is not None:
        progress_callback(len(emitted), processed_existing_insert_statements)
    return emitted


def normalize_create_table_statement(statement: str) -> str:
    normalized = statement
    normalized = re.sub(
        r"(?i)\bCHARACTER\s+SET\s*=?\s*\w+",
        "CHARACTER SET utf8mb4",
        normalized,
    )
    normalized = re.sub(
        r"(?i)\bCOLLATE\s*=?\s*\w+",
        "COLLATE utf8mb4_general_ci",
        normalized,
    )
    normalized = _INTEGER_AND_DATE_TYPE_RE.sub(
        lambda match: re.sub(r"\s*\(\s*\d+\s*\)", "", match.group(0)),
        normalized,
    )
    normalized = _CURRENT_TIMESTAMP_PRECISION_RE.sub(
        lambda match: match.group(0).split("(", 1)[0],
        normalized,
    )
    return normalized


def build_struct_output(struct_statements: Sequence[str]) -> str:
    lines: List[str] = ["SET FOREIGN_KEY_CHECKS=0;"]
    for statement in struct_statements:
        normalized_statement = statement.strip()
        if normalized_statement.upper().startswith("CREATE TABLE"):
            normalized_statement = normalize_create_table_statement(normalized_statement)
        lines.append(f"{normalized_statement};")
    lines.append("SET FOREIGN_KEY_CHECKS=1;")
    lines.append("")
    return "\n".join(lines)


def build_data_output(insert_statements: Sequence[str], table_name: str) -> str:
    lines: List[str] = ["SET FOREIGN_KEY_CHECKS=0;"]
    lines.extend(stmt.strip() for stmt in insert_statements)
    lines.append("COMMIT;")
    lines.append("SET FOREIGN_KEY_CHECKS=1;")
    lines.append(f"SELECT COUNT(*) AS total_rows FROM {table_name};")
    lines.append("")
    return "\n".join(lines)

def classify_ti(has_struct: bool, has_insert: bool) -> str:
    if has_struct and has_insert:
        return "TI"
    if has_struct:
        return "T"
    if has_insert:
        return "I"
    return ""


def read_sql_text(path: Path) -> str:
    return path.read_text(encoding="utf-8-sig", errors="replace")


def build_log_content(
    run_dt: datetime,
    active_config_path: Path,
    config_values: Dict[str, str],
    results: Sequence[FileResult],
) -> str:
    lines: List[str] = []
    normalized = normalize_config_values(config_values)
    lines.append(f"RunDateTime\t{run_dt.isoformat(sep=' ', timespec='seconds')}")
    lines.append(f"Program\t{PROGRAM_NAME}")
    lines.append(f"ConfigFile\t{active_config_path}")
    lines.append("InputParameters")
    for key in CONFIG_KEYS:
        lines.append(f"{key}\t{normalized.get(key, '')}")
    lines.append("")
    lines.append("Filename\tTI\ttotrow\tnew totrow\tduration\tstart\tend\tstatus")

    for item in results:
        status_text = item.status if item.status == "S" else f"F - {item.error}".replace("\t", " ").replace("\n", " ")
        lines.append(
            "\t".join(
                [
                    item.path.name,
                    item.ti,
                    number_for_log(item.original_insert_count),
                    number_for_log(item.new_insert_count),
                    duration_for_log(item.duration_seconds),
                    time_for_log(item.start_dt),
                    time_for_log(item.end_dt),
                    status_text,
                ]
            )
        )

    lines.append("")
    lines.append(f"SuccessCount\t{sum(1 for item in results if item.status == 'S')}")
    lines.append(f"FailCount\t{sum(1 for item in results if item.status == 'F')}")
    lines.append("")
    return "\n".join(lines)


def db_fields_complete(config: Dict[str, str]) -> bool:
    return all(
        normalize_config_values(config).get(key, "").strip()
        for key in ["db_host", "db_user", "db_pw", "db_name", "db_port"]
    )


def informational_db_precheck(config: Dict[str, str]) -> Tuple[bool, str]:
    normalized = normalize_config_values(config)
    host = normalized["db_host"].strip()
    port = parse_db_port(normalized["db_port"])

    try:
        with socket.create_connection((host, port), timeout=5) as sock:
            sock.settimeout(5)
            payload = sock.recv(64)
            if payload:
                return True, f"DB connection precheck succeeded for {host}:{port}."
            return True, f"DB port {host}:{port} accepted the connection."
    except Exception as exc:
        return False, f"DB connection precheck failed for {host}:{port}: {exc}"


def mysql_connection_test(config: Dict[str, str]) -> Tuple[bool, str]:
    normalized = normalize_config_values(config)
    host = normalized["db_host"].strip()
    user = normalized["db_user"].strip()
    password = normalized["db_pw"]
    database = normalized["db_name"].strip()
    port = parse_db_port(normalized["db_port"])

    try:
        import pymysql  # type: ignore

        connection = pymysql.connect(
            host=host,
            user=user,
            password=password,
            database=database,
            port=port,
            connect_timeout=5,
            read_timeout=5,
            write_timeout=5,
        )
        connection.close()
        return True, f"MySQL DB connection is successful for {host}:{port}/{database}."
    except ModuleNotFoundError:
        pass
    except Exception as exc:
        return False, f"MySQL DB connection failed for {host}:{port}/{database}: {exc}"

    try:
        import mysql.connector  # type: ignore

        connection = mysql.connector.connect(
            host=host,
            user=user,
            password=password,
            database=database,
            port=port,
            connection_timeout=5,
        )
        connection.close()
        return True, f"MySQL DB connection is successful for {host}:{port}/{database}."
    except ModuleNotFoundError:
        pass
    except Exception as exc:
        return False, f"MySQL DB connection failed for {host}:{port}/{database}: {exc}"

    mysql_binary = shutil.which("mysql")
    if mysql_binary:
        env = dict(__import__("os").environ)
        env["MYSQL_PWD"] = password
        try:
            completed = subprocess.run(
                [
                    mysql_binary,
                    "-h", host,
                    "-P", str(port),
                    "-u", user,
                    "-D", database,
                    "-e", "SELECT 1",
                ],
                capture_output=True,
                text=True,
                env=env,
                timeout=10,
                check=False,
            )
        except Exception as exc:
            return False, f"MySQL DB connection failed for {host}:{port}/{database}: {exc}"
        if completed.returncode == 0:
            return True, f"MySQL DB connection is successful for {host}:{port}/{database}."
        stderr_text = completed.stderr.strip() or completed.stdout.strip() or "mysql command returned a non-zero exit status."
        return False, f"MySQL DB connection failed for {host}:{port}/{database}: {stderr_text}"

    ok, message = informational_db_precheck(config)
    if ok:
        return False, message + " Full credential validation could not be completed because no MySQL Python client or mysql CLI was found."
    return False, message


def prompt_for_specific_config_values(config: Dict[str, str], keys: Sequence[str]) -> Dict[str, str]:
    updated = dict(normalize_config_values(config))
    for key in keys:
        current_value = updated.get(key, "")
        entered_value = prompt_text(f"{PROMPT_LABELS[key]} [{current_value}]: ").strip()
        if entered_value:
            updated[key] = entered_value
    return updated


def run_batch_db_connection_flow(config: Dict[str, str], active_config_path: Path) -> Dict[str, str]:
    current = dict(normalize_config_values(config))
    db_keys = ["db_host", "db_user", "db_pw", "db_port", "db_name"]

    while True:
        if not db_fields_complete(current):
            print("MySQL DB parameter must be complete to continue testing.")
            if prompt_yes_no("Bypass MySQL DB connection testing? [Y/N]: "):
                print("MySQL DB connection testing will not be performed and user must update the connection string in the generated Batch Upload Script file.")
                return current
            current = prompt_for_specific_config_values(current, db_keys)
            current["db_port"] = str(parse_db_port(current.get("db_port")))
            write_config_file(active_config_path, current)
            continue

        ok, message = mysql_connection_test(current)
        print(message)
        if ok:
            print("MySQL DB connection is successful and can proceed to generate Batch Upload Script File.")
            write_config_file(active_config_path, current)
            return current

        if prompt_yes_no("Change the MySQL DB parameters? [Y/N]: "):
            current = prompt_for_specific_config_values(current, db_keys)
            current["db_port"] = str(parse_db_port(current.get("db_port")))
            write_config_file(active_config_path, current)
            continue

        print("MySQL DB connection testing will not be performed and user must update the connection string in the generated Batch Upload Script file.")
        return current


def resolve_generated_insert_paths(input_expr: str, insert_prefix: str) -> List[Path]:
    input_path = Path(input_expr)
    if input_path.exists():
        if input_path.is_file():
            files = [input_path]
        elif input_path.is_dir():
            files = [p for p in input_path.iterdir() if p.is_file()]
        else:
            files = []
    else:
        files = [Path(p) for p in glob.glob(input_expr)]

    filtered = [
        p for p in files
        if p.is_file() and p.suffix.lower() == ".sql" and p.name.startswith(insert_prefix)
    ]
    return sorted(filtered, key=lambda p: (p.name.lower(), str(p).lower()))


def determine_output_dir_from_input(input_expr: str) -> Path:
    input_path = Path(input_expr)
    if input_path.exists():
        if input_path.is_dir():
            return input_path.resolve()
        if input_path.is_file():
            return input_path.resolve().parent
    wildcard_parent = Path(input_expr).expanduser().parent
    if str(wildcard_parent) not in {"", "."}:
        return wildcard_parent.resolve()
    return Path.cwd()


def shell_quote(value: str) -> str:
    return shlex.quote(value)


def build_restore_script_content(sql_files: Sequence[Path], config: Dict[str, str], error_log_name: str) -> str:
    normalized = normalize_config_values(config)
    lines: List[str] = [
        "#!/usr/bin/env bash",
        "set -u",
        "",
        f"DB_HOST={shell_quote(normalized.get('db_host', '').strip())}",
        f"DB_USER={shell_quote(normalized.get('db_user', '').strip())}",
        f"DB_PW={shell_quote(normalized.get('db_pw', ''))}",
        f"DB_PORT={shell_quote(str(parse_db_port(normalized.get('db_port'))))}",
        f"DB_NAME={shell_quote(normalized.get('db_name', '').strip())}",
        f"ERR_LOG={shell_quote(error_log_name)}",
        "",
        "timestamp_now() {",
        "  local main_part nano_part ampm_part",
        "  main_part=$(date '+%Y-%b-%d %I:%M:%S')",
        "  nano_part=$(date '+%N')",
        "  ampm_part=$(date '+%p' | tr '[:upper:]' '[:lower:]')",
        "  printf '%s%s%s' \"$main_part\" \"${nano_part:0:4}\" \"$ampm_part\"",
        "}",
        "",
        "format_duration() {",
        "  local total_centiseconds=$1",
        "  local hours=$((total_centiseconds / 360000))",
        "  local minutes=$(((total_centiseconds % 360000) / 6000))",
        "  local centiseconds=$((total_centiseconds % 6000))",
        "  printf '%02d.%02d.%04d' \"$hours\" \"$minutes\" \"$centiseconds\"",
        "}",
        "",
        ': > "$ERR_LOG"',
        "",
    ]

    for sql_file in sql_files:
        abs_path = sql_file.resolve()
        lines.extend([
            f"echo 'Processing file: {abs_path.name}'",
            "start_ns=$(date +%s%N)",
            "start_label=$(timestamp_now)",
            f'if MYSQL_PWD="$DB_PW" mysql -h "$DB_HOST" -P "$DB_PORT" -u "$DB_USER" "$DB_NAME" < {shell_quote(str(abs_path))} 2>>"$ERR_LOG"; then',
            "  status_label=SUCCESS",
            "else",
            "  status_label=FAILED",
            "fi",
            "end_ns=$(date +%s%N)",
            "end_label=$(timestamp_now)",
            "duration_cs=$(((end_ns - start_ns) / 10000000))",
            'duration_label=$(format_duration "$duration_cs")',
            'echo "Start: $start_label | End: $end_label | Duration: $duration_label | Status: $status_label"',
            "echo",
        ])

    lines.extend([
        'echo "Error log: $ERR_LOG"',
        "",
    ])
    return "\n".join(lines)

def timestamp_for_script_log(dt: Optional[datetime] = None) -> str:
    dt = dt or datetime.now()
    return dt.strftime("%Y%m%d_%I%M%S%p").lower()


def generate_sql_batch_upload_script(sql_files: Sequence[Path], config: Dict[str, str], output_dir: Path) -> Tuple[Path, Path]:
    run_dt = datetime.now()
    script_path = output_dir / f"restore_mysql_{timestamp_for_filename(run_dt)}.sh"
    error_log_path = output_dir / f"restore_err_{timestamp_for_script_log(run_dt)}.log"
    script_path.write_text(
        build_restore_script_content(sql_files, config, error_log_path.name),
        encoding="utf-8",
        newline="\n",
    )
    script_path.chmod(0o755)
    return script_path, error_log_path


def display_generated_sql_summary(sql_files: Sequence[Path], outputs: Sequence[Path]) -> None:
    print()
    print("Display Summary")
    print("Input SQL File")
    for sql_file in sql_files:
        print(f"{sql_file.name} - {format_size(sql_file)}")
    if outputs:
        print("Output")
        for output_path in outputs:
            print(f"{output_path.name} - pending generation/use")
    print()


def print_no_generated_insert_recommendations() -> None:
    print("There are no generated SQL files with insert to be included in SQL Batch Upload Script File.")
    print("Run the program first with no -proc input to generate the SQL Files with insert")
    print("If generated SQL files with insert exist but were not detected, run the program first with the following input:")
    print("specify folder or directory of generated SQL Files with insert statement")
    print("-proc 2 to run SQL Batch Upload Script File only")
    print("-insert_prefix=[INSERT_PREFIX], -inspref=[INSERT_PREFIX] to specify the prefix name on generated SQL Files with insert")


def convert_log_to_excel(log_path: Path) -> Path:
    output_path = log_path.with_suffix(".xlsx")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "convertnavbak"

    lines = log_path.read_text(encoding="utf-8", errors="replace").splitlines()
    for row_index, line in enumerate(lines, start=1):
        cells = line.split("	") if "	" in line else [line]
        for col_index, cell in enumerate(cells, start=1):
            worksheet.cell(row=row_index, column=col_index, value=cell)

    table_header_index = None
    for index, line in enumerate(lines, start=1):
        if line.startswith("Filename	"):
            table_header_index = index
            break

    if table_header_index is not None:
        headers = [worksheet.cell(row=table_header_index, column=col).value or "" for col in range(1, worksheet.max_column + 1)]
        for col_index, header in enumerate(headers, start=1):
            worksheet.cell(row=table_header_index, column=col_index).font = Font(bold=True)
            if str(header).strip().lower() in {"totrow", "new totrow"}:
                for row_index in range(table_header_index + 1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_index, column=col_index)
                    if cell.value in {None, ""}:
                        continue
                    try:
                        cell.value = int(str(cell.value).replace(",", ""))
                        cell.number_format = "#,##0"
                    except ValueError:
                        pass
        worksheet.freeze_panes = worksheet.cell(row=table_header_index, column=1)

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(12, max_length + 2), 60)

    workbook.save(output_path)
    return output_path

def process_file(
    path: Path,
    one_struct: bool,
    tot_rows: int,
    insert_prefix: str,
    combined_struct_accumulator: List[str],
) -> FileResult:
    result = FileResult(path=path)
    result.start_dt = datetime.now()
    start_perf = time.perf_counter()

    try:
        display_status_line(format_status_line(path.name, 0, 0, 0), done=False)

        text = read_sql_text(path)
        validation_errors = validate_sql_file_header(text, path)
        if validation_errors:
            raise SpecError("; ".join(validation_errors))

        statements = split_sql_statements(text)
        struct_statements = [stmt for stmt in statements if STRUCT_RE.match(stmt)]
        insert_statements = [stmt for stmt in statements if INSERT_RE.match(stmt)]

        has_struct = bool(struct_statements)
        has_insert = bool(insert_statements)
        result.ti = classify_ti(has_struct, has_insert)
        result.original_insert_count = len(insert_statements)

        if not has_insert and not ANY_INSERT_RE.search(text):
            raise SpecError("absence of insert statement is not valid")
        if not has_insert:
            raise SpecError("absence of insert statement is not valid")

        display_status_line(
            format_status_line(path.name, result.original_insert_count, 0, 0),
            done=False,
        )

        if has_struct:
            if one_struct:
                combined_struct_accumulator.extend(struct_statements)
            else:
                struct_output_path = path.with_name(f"struct_{path.name}")
                struct_output_path.write_text(
                    build_struct_output(struct_statements),
                    encoding="utf-8",
                    newline="\n",
                )
                result.struct_outputs.append(struct_output_path)

        def progress_callback(
            current_new_insert_total: int,
            processed_existing_insert_total: int,
        ) -> None:
            display_status_line(
                format_status_line(
                    path.name,
                    result.original_insert_count,
                    current_new_insert_total,
                    processed_existing_insert_total,
                ),
                done=False,
            )

        table_name = detect_single_insert_target(insert_statements)
        regrouped_inserts = regroup_insert_statements(
            insert_statements,
            tot_rows,
            progress_callback=progress_callback,
        )
        result.new_insert_count = len(regrouped_inserts)
        data_output_path = path.with_name(f"{insert_prefix}{path.name}")
        data_output_path.write_text(
            build_data_output(regrouped_inserts, table_name),
            encoding="utf-8",
            newline="\n",
        )
        result.data_output = data_output_path

        result.status = "S"
        display_status_line(
            format_status_line(
                path.name,
                result.original_insert_count,
                result.new_insert_count,
                result.original_insert_count,
            ),
            done=True,
        )

    except Exception as exc:
        result.status = "F"
        result.error = str(exc)
        display_status_line(
            format_status_line(
                path.name,
                result.original_insert_count,
                result.new_insert_count,
                0,
                error=result.error,
            ),
            done=True,
        )

    finally:
        result.end_dt = datetime.now()
        result.duration_seconds = time.perf_counter() - start_perf

    return result


def format_size(path: Path) -> str:
    size = path.stat().st_size
    if size >= 1024 * 1024:
        return f"{size / (1024 * 1024):.2f} MB"
    if size >= 1024:
        return f"{size / 1024:.2f} KB"
    return f"{size} B"


def display_summary(results: Sequence[FileResult], combined_struct_path: Optional[Path]) -> None:
    print()
    print("Display Summary")
    for item in results:
        print("Input SQL File")
        print(f"{item.path.name} - {format_size(item.path)}")
        print("Output")
        outputs: List[Path] = []
        outputs.extend(item.struct_outputs)
        if item.data_output is not None:
            outputs.append(item.data_output)
        if outputs:
            for output_path in outputs:
                print(f"{output_path.name} - {format_size(output_path)}")
        else:
            status_text = item.status if item.status == "S" else f"F - {item.error}"
            print(status_text)
        print()

    if combined_struct_path is not None:
        print("Output")
        print(f"{combined_struct_path.name} - {format_size(combined_struct_path)}")
        print()


def display_log_file(log_path: Path) -> None:
    print()
    print("Log File Content")
    lines = log_path.read_text(encoding="utf-8", errors="replace").splitlines()
    table_header_index = None
    for index, line in enumerate(lines):
        if line.startswith("Filename	"):
            table_header_index = index
            break

    if table_header_index is None:
        print(log_path.read_text(encoding="utf-8", errors="replace"), end="")
        return

    for line in lines[:table_header_index]:
        if line:
            print(line)
        else:
            print()

    table_lines = []
    for line in lines[table_header_index:]:
        if not line:
            break
        table_lines.append(line.split("	"))

    widths = [0] * max(len(row) for row in table_lines)
    for row in table_lines:
        for idx, cell in enumerate(row):
            widths[idx] = max(widths[idx], len(cell))

    header_lookup = {idx: cell.strip().lower() for idx, cell in enumerate(table_lines[0])}
    for row_index, row in enumerate(table_lines):
        padded: List[str] = []
        for idx, cell in enumerate(row):
            if row_index > 0 and header_lookup.get(idx, "") in {"totrow", "new totrow"}:
                padded.append(cell.rjust(widths[idx]))
            else:
                padded.append(cell.ljust(widths[idx]))
        print("  ".join(padded))

    remaining = lines[table_header_index + len(table_lines):]
    if remaining:
        print()
        for line in remaining:
            print(line)

def main(argv: Optional[Sequence[str]] = None) -> int:
    argv = list(argv if argv is not None else sys.argv[1:])
    parser = build_parser()
    args = parser.parse_args(argv)

    if args.show_help:
        print_examples(parser)
        return 0

    try:
        active_config_path, initial_config = resolve_active_config(args)
    except SpecError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        print_examples(parser)
        return 1

    try:
        config = review_and_confirm_config(active_config_path, initial_config)
        input_value = config.get("input", "").strip()
        if not input_value:
            raise SpecError("input is required")

        proc = parse_proc(config.get("proc"))
        one_struct = parse_one_struct(config.get("one_struct", ""))
        tot_rows = parse_tot_rows(config.get("tot_rows"))
        insert_prefix = parse_insert_prefix(config.get("insert_prefix"))
        db_port = parse_db_port(config.get("db_port"))
        config["proc"] = str(proc)
        config["db_port"] = str(db_port)
        write_config_file(active_config_path, config)
    except SpecError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        print_examples(parser)
        return 1

    if proc in {0, 1}:
        if db_fields_complete(config):
            ok, message = informational_db_precheck(config)
            print(message)
            if not ok:
                print("This will be tested again before generating SQL Batch Upload Script.")
        else:
            print(
                "DB host, user, password, database name, and port are incomplete. "
                "These will be asked again before generating SQL Batch Upload Script."
            )

        run_dt = datetime.now()
        files = resolve_input_paths(input_value)
        output_dir = determine_shared_output_dir(files)
        log_path = output_dir / "convertnavbak.log"
        rotate_existing_log(log_path, run_dt)

        if not files:
            log_content = build_log_content(
                run_dt=run_dt,
                active_config_path=active_config_path,
                config_values=config,
                results=[],
            )
            log_path.write_text(log_content, encoding="utf-8", newline="\n")
            print(f"No matching input files found. Log written to: {log_path}", file=sys.stderr)
            return 1

        combined_struct_statements: List[str] = []
        results: List[FileResult] = []
        had_success = False

        for path in files:
            result = process_file(
                path=path,
                one_struct=one_struct,
                tot_rows=tot_rows,
                insert_prefix=insert_prefix,
                combined_struct_accumulator=combined_struct_statements,
            )
            if result.status == "S":
                had_success = True
            results.append(result)

        combined_struct_path: Optional[Path] = None
        if one_struct and combined_struct_statements:
            combined_struct_path = output_dir / f"struct_{timestamp_for_filename(run_dt)}.sql"
            combined_struct_path.write_text(
                build_struct_output(combined_struct_statements),
                encoding="utf-8",
                newline="\n",
            )

        log_content = build_log_content(
            run_dt=run_dt,
            active_config_path=active_config_path,
            config_values=config,
            results=results,
        )
        log_path.write_text(log_content, encoding="utf-8", newline="\n")

        print(f"Log written to: {log_path}")
        if combined_struct_path is not None:
            print(f"Combined struct written to: {combined_struct_path}")

        display_summary(results, combined_struct_path)

        if prompt_yes_no("Display the content of the log file? [Y/N]: "):
            display_log_file(log_path)

        if prompt_yes_no("Convert convertnavbak.log to excel file? [Y/N]: "):
            excel_path = convert_log_to_excel(log_path)
            print(f"Excel file written to: {excel_path}")

        if proc == 1:
            return 0 if had_success else 1

        if not prompt_yes_no("Generate SQL Batch Upload Script File? [Y/N]: "):
            return 0 if had_success else 1

        if prompt_yes_no("Test MySQL DB connection? [Y/N]: "):
            config = run_batch_db_connection_flow(config, active_config_path)
        else:
            print("MySQL DB connection testing will not be performed and user must update the connection string in the generated Batch Upload Script file.")

        generated_insert_files = [item.data_output for item in results if item.status == "S" and item.data_output is not None]
        if not generated_insert_files:
            print_no_generated_insert_recommendations()
            return 1

        script_path, error_log_path = generate_sql_batch_upload_script(generated_insert_files, config, output_dir)
        print(f"SQL Batch Upload Script written to: {script_path}")
        print(f"Restore error log path configured as: {error_log_path}")
        return 0 if had_success else 1

    output_dir = determine_output_dir_from_input(input_value)
    generated_sql_files = resolve_generated_insert_paths(input_value, insert_prefix)
    if not generated_sql_files:
        print_no_generated_insert_recommendations()
        return 1

    preview_script_path = output_dir / f"restore_mysql_{timestamp_for_filename(datetime.now())}.sh"
    display_generated_sql_summary(generated_sql_files, [preview_script_path])
    config = run_batch_db_connection_flow(config, active_config_path)
    script_path, error_log_path = generate_sql_batch_upload_script(generated_sql_files, config, output_dir)
    print(f"SQL Batch Upload Script written to: {script_path}")
    print(f"Restore error log path configured as: {error_log_path}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
